"""
Проверка классификации обращений моделью и подсчёт по темам.

Идея:
1. По листу «исх» построчно: LLM решает, соответствует ли «Текст из источника
   обратной связи» присвоенному «Кластер сигналов 2 уровня» -> True/False в новую
   колонку (verdict_col).
2. Считаем число True по каждому классу (кластеру).
3. В лист «динамика» (классы в колонке «тема») в новую колонку «кол-во от модели»
   пишем это число напротив соответствующей темы.

Файл перезаписывается через openpyxl, поэтому исходное оформление и шапки (включая
пустую первую строку) сохраняются — добавляются только новые колонки.
"""
from __future__ import annotations

import json
import logging
import re
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from .overview_reader import (
    _find_col, _match_key, _norm, _pick_sheet, _read_sheet, _similarity,
)

logger = logging.getLogger(__name__)

_ISH_HEADER = ["источник", "кластер", "текст", "объект", "сигнал", "дата"]
_DYN_HEADER = ["продукт", "тема", "динамик", "недел", "сигнал", "кол-во"]


# ---------------------------------------------------------------- LLM-обёртка
def _ask_llm(llm, prompt: str) -> str:
    """Унифицированный вызов: LangChain-модель (.invoke) или обычный callable."""
    if hasattr(llm, "invoke"):
        out = llm.invoke(prompt)
        return getattr(out, "content", str(out))
    if callable(llm):
        return str(llm(prompt))
    raise TypeError("llm должен быть LangChain-моделью с .invoke или callable(str)->str")


def _parse_bool_array(text: str, n: int) -> List[bool]:
    """Достаём JSON-массив true/false длины n; недостающее -> True (не теряем сигнал)."""
    s = text.strip().replace("```json", "").replace("```", "")
    m = re.search(r"\[.*\]", s, re.DOTALL)
    vals: List[bool] = []
    if m:
        try:
            arr = json.loads(m.group(0))
            for v in arr:
                if isinstance(v, bool):
                    vals.append(v)
                elif isinstance(v, (int, float)):
                    vals.append(bool(v))
                elif isinstance(v, str):
                    vals.append(v.strip().lower() in ("true", "1", "да", "yes", "y"))
        except Exception:
            pass
    if len(vals) < n:          # на всякий случай добиваем True
        vals += [True] * (n - len(vals))
    return vals[:n]


def _verify_batch(llm, items: List[Tuple[str, str]]) -> List[bool]:
    """items: список (текст, класс). Возвращает список True/False той же длины."""
    lines = []
    for i, (text, cluster) in enumerate(items, 1):
        t = " ".join(str(text).split())[:400]
        c = " ".join(str(cluster).split())[:160]
        lines.append(f'{i}. Класс: "{c}"\n   Текст: "{t}"')
    prompt = (
        "Ты проверяешь корректность классификации обращений пользователей.\n"
        "Для КАЖДОГО пункта определи, соответствует ли текст обратной связи "
        "указанному классу (теме) по смыслу.\n"
        "Ответь СТРОГО JSON-массивом из true/false той же длины и в том же порядке, "
        "без пояснений и текста вокруг. Пример: [true, false, true]\n\n"
        + "\n".join(lines)
    )
    try:
        raw = _ask_llm(llm, prompt)
        return _parse_bool_array(raw, len(items))
    except Exception as e:                       # модель упала — не теряем строки
        logger.warning("Проверка батча не удалась (%s), считаю все True", e)
        return [True] * len(items)


# ---------------------------------------------------------- поиск шапки в openpyxl
def _header_row_ws(ws, keywords: List[str], max_scan: int = 10) -> int:
    """1-индексный номер строки заголовков в листе openpyxl."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        joined = " ".join(
            _norm(c.value) for c in ws[r] if c.value is not None
        )
        if sum(1 for kw in keywords if kw in joined) >= 2:
            return r
    return 1


def _col_index_ws(ws, header_row: int, predicate) -> Optional[int]:
    for c in ws[header_row]:
        if c.value is not None and predicate(_norm(c.value)):
            return c.column
    return None


# ----------------------------------------------------------------------- main
def verify_and_count(
    xlsx_path: str,
    llm,
    output_path: Optional[str] = None,
    batch_size: int = 15,
    verdict_col: str = "Соответствие модели",
    count_col: str = "кол-во от модели",
    mapping: Optional[dict] = None,
) -> Tuple[str, dict]:
    """
    Проверяет классификацию построчно, пишет True/False в «исх» и число True по
    каждой теме в «динамика» (новая колонка count_col). Возвращает (путь, отчёт).
    """
    mapping = mapping or {}
    output_path = output_path or xlsx_path
    xls = pd.ExcelFile(xlsx_path)

    ish_sheet = mapping.get("_src_sheet") or _pick_sheet(xls, ["исх", "source", "сигнал", "данны"]) or xls.sheet_names[0]
    dyn_sheet = mapping.get("_dyn_sheet") or _pick_sheet(xls, ["динамик", "dynamic", "процент"])
    if not dyn_sheet:
        raise ValueError("Не нашёл лист «динамика».")

    df = _read_sheet(xls, ish_sheet, _ISH_HEADER)
    c_text = mapping.get("text") or _find_col(df, ["текст из источник", "текст обратн", "текст", "сообщ", "обращени"])
    c_l2 = mapping.get("cluster_l2") or _find_col(df, ["кластер сигналов 2", "кластер 2", "кластер второго", "тема", "класс"])
    if c_text is None or c_l2 is None:
        raise ValueError(
            f"В «{ish_sheet}» не нашёл колонки текста/кластера. Колонки: {list(df.columns)}"
        )

    # --- построчная проверка моделью (батчами) ---
    rows = [(df.iloc[i][c_text], df.iloc[i][c_l2]) for i in range(len(df))]
    verdicts: List[bool] = []
    for start in range(0, len(rows), batch_size):
        verdicts += _verify_batch(llm, rows[start:start + batch_size])
    verdicts = verdicts[:len(rows)]

    # --- подсчёт True по классам (по кластеру 2 уровня) ---
    counts_by_key: Dict[str, int] = {}
    for (text, l2), ok in zip(rows, verdicts):
        if ok and pd.notna(l2) and str(l2).strip():
            k = _match_key(l2)
            counts_by_key[k] = counts_by_key.get(k, 0) + 1

    # --- сопоставление с темами из «динамика» ---
    df_dyn = _read_sheet(xls, dyn_sheet, _DYN_HEADER)
    theme_col = mapping.get("dyn_key") or _find_col(df_dyn, ["тема", "класс", "проблема"])
    if theme_col is None:
        raise ValueError(f"В «{dyn_sheet}» не нашёл колонку «тема». Колонки: {list(df_dyn.columns)}")

    def _count_for(theme) -> int:
        mk = _match_key(theme)
        if mk in counts_by_key:
            return counts_by_key[mk]
        for k, v in counts_by_key.items():               # по вхождению
            if len(k) >= 10 and len(mk) >= 10 and (mk in k or k in mk):
                return v
        best_v, best_sc = 0, 0.0                          # по схожести
        for k, v in counts_by_key.items():
            sc = _similarity(theme, k)
            if sc > best_sc:
                best_v, best_sc = v, sc
        return best_v if best_sc >= 0.5 else 0

    theme_counts = {str(t): _count_for(t) for t in df_dyn[theme_col].dropna()}

    # --- запись обратно через openpyxl (сохраняем оформление и пустые строки) ---
    wb = load_workbook(xlsx_path)
    ws_ish = wb[ish_sheet]
    hr = _header_row_ws(ws_ish, _ISH_HEADER)
    # колонка вердикта — добавляем справа
    vcol = ws_ish.max_column + 1
    ws_ish.cell(row=hr, column=vcol, value=verdict_col)
    for i, ok in enumerate(verdicts):
        ws_ish.cell(row=hr + 1 + i, column=vcol, value=bool(ok))

    ws_dyn = wb[dyn_sheet]
    hrd = _header_row_ws(ws_dyn, _DYN_HEADER)
    tcol = _col_index_ws(ws_dyn, hrd, lambda n: ("тема" in n or "класс" in n or "проблема" in n))
    ccol = ws_dyn.max_column + 1
    ws_dyn.cell(row=hrd, column=ccol, value=count_col)
    if tcol:
        for r in range(hrd + 1, ws_dyn.max_row + 1):
            tv = ws_dyn.cell(row=r, column=tcol).value
            if tv is not None and str(tv).strip():
                ws_dyn.cell(row=r, column=ccol, value=int(theme_counts.get(str(tv), _count_for(tv))))

    try:
        wb.save(output_path)
        saved_to = output_path
    except (PermissionError, OSError) as e:
        # файл, скорее всего, открыт в Excel/R7 — сохраняем копию рядом
        import os
        base, ext = os.path.splitext(output_path)
        saved_to = base + "_verified" + (ext or ".xlsx")
        wb.save(saved_to)
        logger.warning("Не удалось записать в %s (%s). Сохранил в %s",
                       output_path, e, saved_to)

    report = {
        "_src_sheet": ish_sheet, "_dyn_sheet": dyn_sheet,
        "text_col": c_text, "cluster_col": c_l2, "theme_col": theme_col,
        "rows_checked": len(rows),
        "true_count": int(sum(1 for v in verdicts if v)),
        "false_count": int(sum(1 for v in verdicts if not v)),
        "per_theme": theme_counts,
        "verdict_col": verdict_col, "count_col": count_col,
        "output": saved_to,
    }
    logger.info("Проверка завершена: %s/%s True, тем: %s",
                report["true_count"], report["rows_checked"], len(theme_counts))
    return saved_to, report


def format_verify_report(report: dict) -> str:
    lines = ["=== Проверка классификации моделью ===",
             f"  лист исх: {report['_src_sheet']}  |  динамика: {report['_dyn_sheet']}",
             f"  колонка текста:   {report['text_col']}",
             f"  колонка кластера: {report['cluster_col']}",
             f"  проверено строк: {report['rows_checked']}  "
             f"(True: {report['true_count']}, False: {report['false_count']})",
             f"  записано: «{report['verdict_col']}» в исх, «{report['count_col']}» в динамику",
             "  кол-во от модели по темам:"]
    for theme, cnt in report["per_theme"].items():
        lines.append(f"    – {theme[:60]}: {cnt}")
    return "\n".join(lines)
